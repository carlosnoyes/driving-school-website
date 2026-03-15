import openpyxl
from datetime import datetime, timedelta
import random

wb = openpyxl.load_workbook('vehicle_tracker/Vehicles_2026.xlsx')

# Unmerge data-area merged cells that would block writing
for name in wb.sheetnames:
    ws = wb[name]
    # Collect ranges to unmerge (can't modify during iteration)
    to_unmerge = []
    for mr in ws.merged_cells.ranges:
        # Keep title/subtitle merges (rows 1-3), unmerge anything in data rows
        if mr.min_row >= 4:
            to_unmerge.append(str(mr))
    for r in to_unmerge:
        ws.unmerge_cells(r)
        print(f'  Unmerged {r} in {name}')

vehicles = [
    'COLDRV1 22 TOY COR',
    'COLDRV2 17 NISSAN ALTIMA',
    'COLDRV3 17 TOY PRIUS',
    'COLDRV4 2015 HONDA ACCORD',
    'COLDRV5 18 HONDA ACCORD',
    'COLDRV8 21 TOY COR',
]

triggers = ['MAINTENANCE', 'REPAIR', 'BREAK DOWN', 'ACCIDENT']
trigger_weights = [60, 25, 10, 5]

categories = [
    'Routine / Wear Items',
    'Engine & Cooling',
    'Transmission & Drivetrain',
    'Steering & Suspension',
    'Electrical & Other Major',
]

shops = [
    'Quick Lube Express', 'Toyota of Columbus', 'Honda Service Center',
    'Pep Boys', 'Firestone', 'Jiffy Lube', 'Midas Auto', 'AAMCO',
    'Discount Tire', 'Meineke',
]

service_details_by_cat = {
    'Routine / Wear Items': [
        ('Oil change + filter', 45, 85),
        ('Brake pad replacement (front)', 180, 320),
        ('Brake pad replacement (rear)', 160, 290),
        ('Tire rotation + balance', 40, 80),
        ('New tires (set of 4)', 450, 800),
        ('Battery replacement', 120, 220),
        ('Wiper blade replacement', 20, 45),
        ('Air filter replacement', 25, 55),
        ('Cabin air filter', 30, 60),
        ('Headlight bulb replacement', 15, 50),
        ('Alignment', 80, 130),
        ('Brake fluid flush', 70, 120),
        ('Coolant flush', 90, 150),
        ('Spark plug replacement', 100, 250),
    ],
    'Engine & Cooling': [
        ('Radiator replacement', 400, 750),
        ('Water pump replacement', 350, 600),
        ('Thermostat replacement', 150, 300),
        ('Coolant leak repair', 200, 450),
        ('Motor mount replacement', 250, 500),
        ('Engine diagnostic', 80, 150),
        ('Valve cover gasket', 200, 400),
    ],
    'Transmission & Drivetrain': [
        ('Transmission fluid change', 150, 300),
        ('CV axle replacement', 250, 500),
        ('Clutch replacement', 800, 1500),
        ('Differential service', 100, 200),
        ('CV boot replacement', 180, 350),
    ],
    'Steering & Suspension': [
        ('Strut replacement (pair)', 400, 800),
        ('Control arm replacement', 250, 500),
        ('Wheel bearing replacement', 200, 450),
        ('Tie rod end replacement', 150, 350),
        ('Power steering fluid flush', 80, 140),
        ('Sway bar link replacement', 100, 250),
    ],
    'Electrical & Other Major': [
        ('Alternator replacement', 350, 650),
        ('Starter motor replacement', 300, 550),
        ('AC recharge', 100, 200),
        ('AC compressor replacement', 500, 900),
        ('O2 sensor replacement', 150, 350),
        ('Catalytic converter replacement', 800, 2000),
        ('Check engine light diagnosis', 80, 150),
    ],
}

base_mileages = {
    'COLDRV1 22 TOY COR': 28000,
    'COLDRV2 17 NISSAN ALTIMA': 78000,
    'COLDRV3 17 TOY PRIUS': 82000,
    'COLDRV4 2015 HONDA ACCORD': 105000,
    'COLDRV5 18 HONDA ACCORD': 62000,
    'COLDRV8 21 TOY COR': 38000,
}

random.seed(42)

# ---- SERVICE LOG ----
ws = wb['Service Log']

entries = []
for v in vehicles:
    mi = base_mileages[v]
    n_services = random.randint(5, 12)
    start_date = datetime(2025, 1, 15)
    for i in range(n_services):
        days_offset = random.randint(20, 90)
        start_date += timedelta(days=days_offset)
        if start_date > datetime(2026, 3, 10):
            break
        mi += random.randint(800, 4000)
        trigger = random.choices(triggers, weights=trigger_weights, k=1)[0]

        if trigger == 'MAINTENANCE':
            cat = random.choices(categories, weights=[70, 10, 5, 10, 5], k=1)[0]
        elif trigger == 'BREAK DOWN':
            cat = random.choices(categories, weights=[10, 30, 25, 15, 20], k=1)[0]
        elif trigger == 'ACCIDENT':
            cat = random.choices(categories, weights=[5, 10, 5, 50, 30], k=1)[0]
        else:
            cat = random.choices(categories, weights=[40, 20, 10, 15, 15], k=1)[0]

        details_list = service_details_by_cat[cat]
        detail_name, cost_lo, cost_hi = random.choice(details_list)
        cost = round(random.uniform(cost_lo, cost_hi), 2)
        shop = random.choice(shops)
        ref = f'INV-{random.randint(10000, 99999)}'
        notes = ''
        if trigger == 'BREAK DOWN':
            notes = random.choice([
                "Vehicle wouldn't start", 'Overheating on highway',
                'Strange noise from engine', 'Stalled at intersection',
                'Warning light came on',
            ])
        elif trigger == 'ACCIDENT':
            notes = random.choice([
                'Minor fender bender', 'Rear-ended at stop light',
                'Parking lot incident', 'Side mirror clipped',
            ])

        entries.append((v, start_date, mi, trigger, notes, detail_name, cat, shop, cost, ref))

entries.sort(key=lambda x: x[1])

for i, (v, dt, mi, trig, notes, detail, cat, shop, cost, ref) in enumerate(entries):
    r = 4 + i
    ws.cell(row=r, column=1, value=v)
    ws.cell(row=r, column=2, value=dt).number_format = 'MM/DD/YYYY'
    ws.cell(row=r, column=3, value=mi)
    ws.cell(row=r, column=4, value=trig)
    ws.cell(row=r, column=5, value=notes)
    ws.cell(row=r, column=6, value=detail)
    ws.cell(row=r, column=7, value=cat)
    ws.cell(row=r, column=8, value=shop)
    ws.cell(row=r, column=9, value=cost)
    ws.cell(row=r, column=10, value=ref)

print(f'Service Log: {len(entries)} entries added')

# ---- FUEL LOG ----
ws = wb['Fuel Log']
fuel_entries = []
for v in vehicles:
    n_fillups = random.randint(10, 20)
    dt = datetime(2025, 1, 10)
    for i in range(n_fillups):
        dt += timedelta(days=random.randint(7, 21))
        if dt > datetime(2026, 3, 10):
            break
        gallons = round(random.uniform(7.5, 14.5), 2)
        if 'PRIUS' in v:
            miles = round(random.uniform(350, 550), 1)
        elif 'ALTIMA' in v or 'ACCORD' in v:
            miles = round(random.uniform(250, 400), 1)
        else:
            miles = round(random.uniform(220, 380), 1)
        cost_per_gal = round(random.uniform(2.85, 3.65), 2)
        fuel_entries.append((v, dt, gallons, miles, cost_per_gal))

fuel_entries.sort(key=lambda x: x[1])

for i, (v, dt, gal, miles, cpg) in enumerate(fuel_entries):
    r = 4 + i
    ws.cell(row=r, column=1, value=v)
    ws.cell(row=r, column=2, value=dt).number_format = 'MM/DD/YYYY'
    ws.cell(row=r, column=3, value=gal)
    ws.cell(row=r, column=4, value=miles)
    ws.cell(row=r, column=5).value = f'=IFERROR(D{r}/C{r},"")'
    ws.cell(row=r, column=6, value=cpg)
    ws.cell(row=r, column=7).value = f'=IFERROR(C{r}*F{r},"")'

print(f'Fuel Log: {len(fuel_entries)} entries added')

# ---- MAJOR SYSTEMS TRACKER ----
ws = wb['Major Systems Tracker']
systems = [
    'Engine', 'Transmission', 'AC System', 'Alternator', 'Radiator',
    'Suspension', 'Steering', 'Catalytic Converter', 'Starter', 'Braking System',
]
event_types = ['Failure', 'Repair', 'Replacement', 'Inspection', 'Preventive']
descriptions = {
    'Engine': ['Misfire detected cyl 3', 'Oil leak at valve cover', 'Check engine light - P0301', 'Engine mount cracked'],
    'Transmission': ['Slipping between 2nd and 3rd', 'Fluid dark and burnt smell', 'Hard shift when cold', 'CV joint clicking'],
    'AC System': ['Not blowing cold', 'Compressor clutch not engaging', 'Refrigerant leak found', 'Blower motor weak'],
    'Alternator': ['Battery not charging', 'Whining noise from belt area', 'Voltage dropping under load'],
    'Radiator': ['Coolant leak at seam', 'Overheating in traffic', 'Plastic end tank cracked'],
    'Suspension': ['Clunking over bumps', 'Strut leaking', 'Uneven tire wear pattern', 'Bouncy ride'],
    'Steering': ['Power steering whine', 'Play in steering wheel', 'Tie rod boot torn'],
    'Catalytic Converter': ['P0420 code - efficiency below threshold', 'Rattle from underneath'],
    'Starter': ['Slow cranking', 'Clicking but not turning over', 'Intermittent no-start'],
    'Braking System': ['Pulsating brake pedal', 'ABS light on', 'Grinding noise', 'Soft brake pedal'],
}

major_entries = []
for v in vehicles:
    if '2015' in v or '17' in v:
        n = random.randint(3, 6)
    elif '18' in v:
        n = random.randint(2, 4)
    else:
        n = random.randint(1, 3)
    dt = datetime(2025, 2, 1)
    for _ in range(n):
        dt += timedelta(days=random.randint(30, 150))
        if dt > datetime(2026, 3, 10):
            break
        sys_name = random.choice(systems)
        evt = random.choice(event_types)
        desc = random.choice(descriptions[sys_name])
        cost = round(random.uniform(150, 1800), 2)
        shop = random.choice(shops)
        warranty = random.choice(['Yes', 'No', 'No', 'No', 'No'])
        warranty_exp = ''
        if warranty == 'Yes':
            warranty_exp = dt + timedelta(days=random.randint(180, 730))
        mi = base_mileages[v] + random.randint(5000, 25000)
        major_entries.append((v, dt, mi, sys_name, evt, desc, cost, shop, warranty, warranty_exp))

major_entries.sort(key=lambda x: x[1])

for i, (v, dt, mi, sys_name, evt, desc, cost, shop, warr, warr_exp) in enumerate(major_entries):
    r = 4 + i
    ws.cell(row=r, column=1, value=v)
    ws.cell(row=r, column=2, value=dt).number_format = 'MM/DD/YYYY'
    ws.cell(row=r, column=3, value=mi)
    ws.cell(row=r, column=4, value=sys_name)
    ws.cell(row=r, column=5, value=evt)
    ws.cell(row=r, column=6, value=desc)
    ws.cell(row=r, column=7, value=cost)
    ws.cell(row=r, column=8, value=shop)
    ws.cell(row=r, column=9, value=warr)
    if warr_exp:
        ws.cell(row=r, column=10, value=warr_exp).number_format = 'MM/DD/YYYY'

print(f'Major Systems Tracker: {len(major_entries)} entries added')

wb.save('vehicle_tracker/Vehicles_2026.xlsx')
print('Saved successfully!')

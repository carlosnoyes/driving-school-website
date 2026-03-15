import openpyxl
from datetime import datetime, timedelta
import random

wb = openpyxl.load_workbook('vehicle_tracker/Vehicles_2026.xlsx')
ws = wb['Major Systems Tracker']

# Find next empty row
next_row = 22  # rows 4-21 are filled

base_mileages = {
    'COLDRV1 22 TOY COR': 28000,
    'COLDRV2 17 NISSAN ALTIMA': 78000,
    'COLDRV3 17 TOY PRIUS': 82000,
    'COLDRV4 2015 HONDA ACCORD': 105000,
    'COLDRV5 18 HONDA ACCORD': 62000,
    'COLDRV8 21 TOY COR': 38000,
}

shops = [
    'Quick Lube Express', 'Toyota of Columbus', 'Honda Service Center',
    'Pep Boys', 'Firestone', 'Jiffy Lube', 'Midas Auto', 'AAMCO',
    'Discount Tire', 'Meineke', 'Goodyear Auto', 'NTB',
]

# Additional entries to flesh out the tracker — focus on vehicles with fewer entries
# and different systems for variety
additional = [
    # COLDRV1 - only had 2 entries, add more
    ('COLDRV1 22 TOY COR', datetime(2025, 7, 14), 34200, 'AC System', 'Failure',
     'Not blowing cold - compressor clutch not engaging', 687.50, 'Toyota of Columbus', 'No', ''),
    ('COLDRV1 22 TOY COR', datetime(2025, 9, 22), 36800, 'Suspension', 'Inspection',
     'Clunking over bumps - front strut mount worn', 185.00, 'Pep Boys', 'No', ''),
    ('COLDRV1 22 TOY COR', datetime(2025, 12, 5), 39500, 'Braking System', 'Repair',
     'Pulsating brake pedal - rotors warped', 445.30, 'Firestone', 'No', ''),
    ('COLDRV1 22 TOY COR', datetime(2026, 2, 10), 41800, 'Starter', 'Inspection',
     'Slow cranking in cold weather - tested OK', 95.00, 'Toyota of Columbus', 'No', ''),

    # COLDRV8 - only had 1 entry, add more
    ('COLDRV8 21 TOY COR', datetime(2025, 6, 18), 44500, 'AC System', 'Repair',
     'Refrigerant leak found at condenser', 520.75, 'Toyota of Columbus', 'Yes',
     datetime(2026, 6, 18)),
    ('COLDRV8 21 TOY COR', datetime(2025, 9, 10), 47200, 'Suspension', 'Replacement',
     'Strut leaking - replaced both front struts', 892.40, 'Meineke', 'No', ''),
    ('COLDRV8 21 TOY COR', datetime(2025, 11, 28), 49800, 'Engine', 'Inspection',
     'Check engine light - P0171 lean condition, cleaned MAF', 165.00, 'Toyota of Columbus', 'No', ''),
    ('COLDRV8 21 TOY COR', datetime(2026, 1, 15), 51300, 'Steering', 'Repair',
     'Tie rod boot torn - replaced outer tie rods', 385.60, 'AAMCO', 'No', ''),

    # COLDRV5 - only had 2 entries, add more
    ('COLDRV5 18 HONDA ACCORD', datetime(2025, 5, 12), 67800, 'Suspension', 'Failure',
     'Control arm bushing cracked - front left', 475.00, 'Honda Service Center', 'No', ''),
    ('COLDRV5 18 HONDA ACCORD', datetime(2025, 8, 25), 71200, 'Engine', 'Repair',
     'Valve cover gasket leaking - oil on exhaust manifold', 380.25, 'Honda Service Center', 'No', ''),
    ('COLDRV5 18 HONDA ACCORD', datetime(2025, 11, 3), 74500, 'Alternator', 'Failure',
     'Battery not charging - alternator output low', 595.80, 'Pep Boys', 'No', ''),
    ('COLDRV5 18 HONDA ACCORD', datetime(2026, 1, 20), 77100, 'Braking System', 'Replacement',
     'ABS light on - wheel speed sensor replaced', 310.45, 'Honda Service Center', 'Yes',
     datetime(2027, 1, 20)),

    # More for COLDRV2 to show aging pattern
    ('COLDRV2 17 NISSAN ALTIMA', datetime(2025, 4, 5), 85600, 'Suspension', 'Failure',
     'Strut leaking - bouncy ride over bumps', 780.00, 'Midas Auto', 'No', ''),
    ('COLDRV2 17 NISSAN ALTIMA', datetime(2025, 10, 15), 96700, 'AC System', 'Repair',
     'Blower motor weak - replaced blower motor resistor', 225.50, 'Meineke', 'No', ''),
    ('COLDRV2 17 NISSAN ALTIMA', datetime(2026, 1, 8), 100200, 'Engine', 'Repair',
     'Oil leak at valve cover - gasket replaced', 345.00, 'Jiffy Lube', 'No', ''),

    # More for COLDRV4 (oldest, should have most issues)
    ('COLDRV4 2015 HONDA ACCORD', datetime(2025, 3, 20), 109800, 'AC System', 'Failure',
     'Compressor clutch not engaging - compressor seized', 1150.00, 'Honda Service Center', 'No', ''),
    ('COLDRV4 2015 HONDA ACCORD', datetime(2025, 6, 10), 112500, 'Steering', 'Repair',
     'Power steering whine - pump leaking, replaced', 625.40, 'AAMCO', 'No', ''),
    ('COLDRV4 2015 HONDA ACCORD', datetime(2025, 9, 8), 116200, 'Engine', 'Replacement',
     'Motor mount cracked - excessive engine vibration', 485.75, 'Honda Service Center', 'No', ''),
    ('COLDRV4 2015 HONDA ACCORD', datetime(2025, 12, 18), 118900, 'Catalytic Converter', 'Inspection',
     'P0420 code - efficiency below threshold, monitoring', 125.00, 'Meineke', 'No', ''),
    ('COLDRV4 2015 HONDA ACCORD', datetime(2026, 2, 25), 121500, 'Starter', 'Failure',
     'Clicking but not turning over - starter replaced', 465.30, 'Honda Service Center', 'No', ''),

    # A couple more for COLDRV3
    ('COLDRV3 17 TOY PRIUS', datetime(2025, 7, 5), 88400, 'Braking System', 'Repair',
     'Soft brake pedal - brake booster vacuum leak', 550.00, 'Toyota of Columbus', 'No', ''),
    ('COLDRV3 17 TOY PRIUS', datetime(2025, 12, 20), 94600, 'Steering', 'Replacement',
     'Wheel bearing replacement - humming noise at speed', 420.15, 'Firestone', 'No', ''),
]

random.seed(99)

for entry in additional:
    v, dt, mi, sys_name, evt, desc, cost, shop, warr, warr_exp = entry
    r = next_row
    ws.cell(row=r, column=1, value=v)
    ws.cell(row=r, column=2, value=dt).number_format = 'MM/DD/YYYY'
    ws.cell(row=r, column=3, value=mi)
    ws.cell(row=r, column=4, value=sys_name)
    ws.cell(row=r, column=5, value=evt)
    ws.cell(row=r, column=6, value=desc)
    ws.cell(row=r, column=7, value=cost)
    ws.cell(row=r, column=8, value=shop)
    ws.cell(row=r, column=9, value=warr)
    if warr_exp:
        ws.cell(row=r, column=10, value=warr_exp).number_format = 'MM/DD/YYYY'
    next_row += 1

print(f'Added {len(additional)} entries (rows 22-{next_row-1})')
print(f'Total Major Systems entries: {next_row - 4}')

wb.save('vehicle_tracker/Vehicles_2026.xlsx')
print('Saved!')
